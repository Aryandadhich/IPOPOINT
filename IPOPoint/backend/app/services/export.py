"""Excel export service — generates an in-memory .xlsx for the user's IPO list."""
from __future__ import annotations
import io

import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment


_HEADERS = [
    "IPO Name", "Open Date", "Close Date", "Allotment Date",
    "Issue Price", "GMP",
    "(S) ACC 1", "Status",
    "(A) ACC 2", "Status",
    "(R) ACC 3", "Status",
    "Total Lots", "Allotment Status", "Shares Allotted",
    "Listing Price", "Listing Gain/Loss", "Notes", "Updated At",
]

_FIELD_ORDER = [
    "name", "open_date", "close_date", "allotment_date",
    "issue_price", "gmp",
    "acc1_applied", "acc1_status",
    "acc2_applied", "acc2_status",
    "acc3_applied", "acc3_status",
    "total_lots", "allotment_status", "shares_allotted",
    "listing_price", "listing_gain", "notes", "updated_at",
]

# Emerald green header fill matching the app theme
_HEADER_FILL  = PatternFill("solid", fgColor="2EA87E")
_HEADER_FONT  = Font(bold=True, color="FFFFFF", size=10)
_HEADER_ALIGN = Alignment(horizontal="center", vertical="center")


def build_excel(rows: list[dict]) -> io.BytesIO:
    """Return a BytesIO containing the Excel workbook for *rows*."""
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "IPO Tracker"

    # Header row
    ws.append(_HEADERS)
    for col_idx, cell in enumerate(ws[1], start=1):
        cell.fill      = _HEADER_FILL
        cell.font      = _HEADER_FONT
        cell.alignment = _HEADER_ALIGN
        # Auto-width approximation
        ws.column_dimensions[cell.column_letter].width = max(
            len(_HEADERS[col_idx - 1]) + 4, 12
        )

    # Data rows
    for r in rows:
        ws.append([r.get(f, "") or "" for f in _FIELD_ORDER])

    # Freeze header
    ws.freeze_panes = "A2"

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf
